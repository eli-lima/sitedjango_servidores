from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, CreateView, FormView, ListView, View
from .models import Ajuda_Custo, DataMajorada, LimiteAjudaCusto, CotaAjudaCusto, MatriculaImportante
from .forms import EnvioDatasForm, ConfirmacaoDatasForm, FiltroRelatorioForm, AdminDatasForm, LimiteAjudaCustoForm, UploadExcelRx2Form, CotaAjudaCustoForm
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Sum
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from openpyxl import Workbook
import pandas as pd
from servidor.models import Servidor
from django.core.paginator import Paginator
from django.contrib.auth.mixins import UserPassesTestMixin
import zipfile
from io import BytesIO
import os
import logging
import requests
import cloudinary.uploader
from django.utils import timezone
from django.http import JsonResponse
from django.utils.timezone import now
from .tasks import process_batch
from celery.result import AsyncResult  # Para lidar com o status da task
from .tasks import process_excel_file  # Task Celery para processamento
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from openpyxl.styles import Alignment
from django.db.models import Sum, Case, When, IntegerField, Count
from django.db import transaction
from seappb.models import Unidade
from .utils import (get_intervalo_mes, calcular_horas_por_unidade, get_limites_horas_por_unidade, \
        calcular_horas_a_adicionar_por_unidade, verificar_limites, get_servidor, get_registros_mes,
                    build_context, datas_adicionar, gerar_e_armazenar_codigo, enviar_email_verificacao,
                    verificar_codigo_expiracao)


from datetime import datetime, timedelta
from weasyprint import HTML
import tempfile
from django.template.loader import get_template
from xhtml2pdf import pisa


# Create your views here.

#dados istribuir horas atualizar
transaction.atomic
def distribuir_horas(request, unidade_id):
    if request.method == 'POST':
        horas = int(request.POST.get('horas'))
        unidade = get_object_or_404(Unidade, id=unidade_id)
        cota = CotaAjudaCusto.objects.filter(unidade=unidade).first()

        if cota and cota.carga_horaria_disponivel >= horas:
            LimiteAjudaCusto.objects.create(unidade=unidade, limite_horas=horas)
            cota.carga_horaria_disponivel -= horas
            cota.save()
            return JsonResponse({'status': 'success', 'message': 'Horas distribuídas com sucesso.'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Carga horária insuficiente.'}, status=400)

    # <-- Aqui está o return faltante
    return JsonResponse({'status': 'error', 'message': 'Método não permitido.'}, status=405)


#relatorio em pdf ajuda_custo
def gerar_pdf_ajuda_custo(request, queryset):
    # Define o caminho do template e o contexto
    template_path = 'relatorio_ajuda_custo_pdf.html'
    context = {'ajudas_custo': queryset}

    # Renderiza o HTML com os dados
    template = get_template(template_path)
    html = template.render(context)

    # Cria a resposta como PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_ajuda_custo.pdf"'

    # Gera o PDF usando xhtml2pdf (pisa)
    pisa_status = pisa.CreatePDF(html, dest=response)

    # Se ocorrer erro na geração do PDF
    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', status=500)

    return response


# upload relatorios do rx2

def upload_excel_rx2(request):
    if request.method == 'POST':
        form = UploadExcelRx2Form(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                # Upload para o Cloudinary
                upload_result = cloudinary.uploader.upload(excel_file, resource_type="raw")
                cloudinary_url = upload_result['url']

                # Envia a tarefa de processamento para o Celery
                task = process_excel_file.delay(cloudinary_url)

                # Informa o usuário e redireciona para a página de status
                messages.success(request, "Arquivo enviado para o Cloudinary e processamento iniciado.")
                return redirect('ajuda_custo:status_task', task_id=task.id)

            except Exception as e:
                messages.error(request, f"Erro ao fazer upload no Cloudinary: {str(e)}")
                return redirect('ajuda_custo:upload_excel_rx2')
    else:
        form = UploadExcelRx2Form()

    return render(request, 'upload_excel_rx2.html', {'form': form})


def status_task(request, task_id):
    task = AsyncResult(task_id)


    # Verifica se a tarefa está pendente
    if task.state == 'PENDING':
        status = "Processamento pendente..."
    # Verifica se a tarefa foi concluída com sucesso
    elif task.state == 'SUCCESS':
        result = task.result
        if result['status'] == 'sucesso':
            status = "Processamento concluído com sucesso!"
        else:
            status = f"Erros encontrados: {', '.join(result['erros'])}"
    # Verifica se houve falha na tarefa
    elif task.state == 'FAILURE':
        status = f"Falha no processamento: {task.result}"
    # Para outros estados
    else:
        status = f"Processamento em andamento... Status: {task.state}"

    return render(request, 'status_task.html', {'status': status})


# def baixar zip arquivos assinados

def criar_arquivo_zip(request, queryset):
    # Cria um buffer de memória para o arquivo zip
    buffer = BytesIO()
    arquivos_adicionados = 0  # Contador para arquivos adicionados
    urls_adicionadas = set()  # Conjunto para rastrear URLs já adicionadas

    try:
        # Cria o arquivo zip em memória
        with zipfile.ZipFile(buffer, 'w') as zip_file:
            for registro in queryset:
                if registro.folha_assinada:
                    try:
                        # Obtém a URL do arquivo armazenado no Cloudinary
                        arquivo_url = registro.folha_assinada.url

                        # Verifica se a URL já foi adicionada
                        if arquivo_url in urls_adicionadas:
                            continue  # Pula para o próximo registro se já foi adicionado

                        logging.info(f"Tentando baixar arquivo: {arquivo_url}")

                        # Baixa o arquivo do Cloudinary
                        arquivo_resposta = requests.get(arquivo_url)
                        if arquivo_resposta.status_code == 200:
                            # Determina a extensão do arquivo a partir do Content-Type
                            content_type = arquivo_resposta.headers.get('Content-Type')
                            if content_type == 'application/pdf':
                                extensao = '.pdf'
                            elif content_type == 'image/jpeg':
                                extensao = '.jpeg'
                            elif content_type == 'image/jpg':
                                extensao = '.jpg'
                            else:
                                extensao = '.pdf'  # Define PDF como padrão se não houver Content-Type

                            # Configura o nome da pasta no ZIP (mês e ano)
                            pasta_nome = registro.data.strftime('%m-%Y')

                            # Configura o nome do arquivo no ZIP
                            nome_arquivo_zip = f"{registro.matricula}_{registro.data.strftime('%Y-%m')}{extensao}"

                            # Caminho completo dentro do ZIP
                            caminho_no_zip = os.path.join(pasta_nome, nome_arquivo_zip)

                            # Adiciona o arquivo ao zip
                            zip_file.writestr(caminho_no_zip, arquivo_resposta.content)
                            arquivos_adicionados += 1  # Incrementa o contador
                            urls_adicionadas.add(arquivo_url)  # Marca a URL como adicionada
                        else:
                            logging.error(f"Erro ao acessar o arquivo {registro.folha_assinada.name} no Cloudinary.")
                            messages.warning(request, f'Arquivo {registro.folha_assinada.name} não pôde ser acessado.')

                    except Exception as e:
                        logging.error(f"Erro ao adicionar arquivo {registro.folha_assinada.name} ao ZIP: {str(e)}")
                        messages.warning(request,
                                         f'Arquivo {registro.folha_assinada.name} não pôde ser adicionado ao ZIP.')

        # Log de quantos arquivos foram adicionados
        logging.info(f"Total de arquivos adicionados ao ZIP: {arquivos_adicionados}")

        # Configura o ponteiro do buffer para o início do arquivo
        buffer.seek(0)

        # Configura o nome do arquivo de download
        filename = "arquivos_assinados.zip"

        # Cria uma resposta HTTP com o tipo de conteúdo para download de arquivos ZIP
        response = HttpResponse(buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    except Exception as e:
        logging.error(f"Erro ao criar o arquivo ZIP: {str(e)}")
        messages.error(request, 'Erro ao criar o arquivo ZIP. Tente novamente mais tarde.')
        return HttpResponse(status=500)  # Retorna um status de erro





#DEF BUSCAR NOME DE SERVIDOR


def buscar_nome_servidor(request):
    matricula = request.GET.get('matricula')
    try:
        servidor = Servidor.objects.get(matricula=matricula)
        data = {
            'nome': servidor.nome,
            'cargo': servidor.cargo
        }
        return JsonResponse(data)
    except Servidor.DoesNotExist:
        return JsonResponse({'error': 'Servidor não encontrado'}, status=404)

#DEF Download Relatorios


def exportar_excel(request):
    # Captura os parâmetros de pesquisa
    query = request.GET.get('query', '')
    data_inicial = request.GET.get('dataInicial')
    data_final = request.GET.get('dataFinal')
    unidade = request.GET.get('unidade')





    # Converte as datas em objetos datetime, se fornecidas
    data_inicial = parse_date(data_inicial) if data_inicial else None
    data_final = parse_date(data_final) if data_final else None


    # Filtrar os dados com base nos parâmetros
    queryset = Ajuda_Custo.objects.all()
    if query:
        queryset = queryset.filter(Q(nome__icontains=query) | Q(matricula__icontains=query))


    if unidade:
        queryset = queryset.filter(unidade=unidade)

    if data_inicial and data_final:
        queryset = queryset.filter(data__range=[data_inicial, data_final])
    elif data_inicial:
        queryset = queryset.filter(data__gte=data_inicial)
    elif data_final:
        queryset = queryset.filter(data__lte=data_final)



    # Processar os resultados em um dicionário
    dados_matriculas = {}
    for item in queryset:
        matricula = item.matricula
        nome = item.nome
        carga_horaria = item.carga_horaria
        data = item.data
        majorado = item.majorado



        if matricula not in dados_matriculas:
            dados_matriculas[matricula] = {'Matrícula': matricula, 'Nome': nome, 'Carga Horária Total': 0,
                                           'Horas Majoradas': 0, 'Horas Normais': 0, 'Total': 0,
                                           'Datas 12 Horas': [], 'Datas 24 Horas': [], '12h': 0, '24h': 0}

        if carga_horaria == '12 horas':
            dados_matriculas[matricula]['Datas 12 Horas'].append(pd.to_datetime(data))
            dados_matriculas[matricula]['Carga Horária Total'] += 12
            if majorado:
                dados_matriculas[matricula]['Horas Majoradas'] += 12
                dados_matriculas[matricula]['12h'] += 1
            else:
                dados_matriculas[matricula]['Horas Normais'] += 12
                dados_matriculas[matricula]['12h'] += 1

        elif carga_horaria == '24 horas':
            dados_matriculas[matricula]['Datas 24 Horas'].append(pd.to_datetime(data))
            dados_matriculas[matricula]['Carga Horária Total'] += 24
            if majorado:
                dados_matriculas[matricula]['Horas Majoradas'] += 24
                dados_matriculas[matricula]['24h'] += 1
            else:
                dados_matriculas[matricula]['Horas Normais'] += 24
                dados_matriculas[matricula]['24h'] += 1



        # Calcular o total de horas
        dados_matriculas[matricula]['Total'] = "{:07d}".format(
            int(f"{dados_matriculas[matricula]['Horas Majoradas']:03d}{dados_matriculas[matricula]['Horas Normais']:03d}")
        )



    # Criar um DataFrame com os dados processados
    df = pd.DataFrame(list(dados_matriculas.values()))

    # Exportar para Excel com openpyxl
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ajuda_custo_resumo.xlsx'

    # Criar uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Ajuda Custo"

    # Adicionar o período de datas no topo da planilha
    ws.merge_cells('A1:J1')
    ws[
        'A1'] = f"Período: {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}" if data_inicial and data_final else 'Período: Todos os Dados'

    # Adicionar os cabeçalhos das colunas
    cabecalhos = ['Matrícula', 'Nome', 'Carga Horária Total', 'Horas Majoradas', 'Horas Normais', 'Total',
                  '12h', '24h', 'Datas 12 Horas', 'Datas 24 Horas']
    ws.append(cabecalhos)

    # Função para formatar as datas em múltiplas linhas
    def formatar_datas_em_linhas(datas):
        linhas = []
        for i in range(0, len(datas), 3):  # Agrupa a cada 3 datas
            linhas.append(', '.join(datas[i:i + 3]))
        return '\n'.join(linhas)  # Junta as linhas com quebra de linha

    # Adicionar os dados do DataFrame na planilha
    for index, row in df.iterrows():
        datas_12_horas = [d.strftime('%d/%m/%Y') for d in row['Datas 12 Horas']]
        datas_24_horas = [d.strftime('%d/%m/%Y') for d in row['Datas 24 Horas']]

        # Formatar as datas em múltiplas linhas
        datas_12_horas_formatadas = formatar_datas_em_linhas(datas_12_horas)
        datas_24_horas_formatadas = formatar_datas_em_linhas(datas_24_horas)

        # Adicionar a linha formatada na planilha
        ws.append([
            row['Matrícula'], row['Nome'], row['Carga Horária Total'],
            row['Horas Majoradas'], row['Horas Normais'], row['Total'],
            row['12h'], row['24h'], datas_12_horas_formatadas, datas_24_horas_formatadas
        ])

    # Ajustar a largura das colunas ignorando a primeira linha
    for col in ws.iter_cols(min_row=2, max_col=ws.max_column):
        max_length = 0
        column_letter = col[0].column_letter  # Obtém a letra da coluna (A, B, C, etc.)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Ajustar o alinhamento e a altura das linhas para as colunas de datas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=10):  # Colunas 9 e 10 são as de datas
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)  # Permite quebra de linha dentro da célula
            num_linhas = str(cell.value).count('\n') + 1  # Conta o número de linhas dentro da célula
            ws.row_dimensions[cell.row].height = 15 * num_linhas  # Ajusta a altura da linha conforme o número de linhas

    # Ajustar o alinhamento central para todas as outras células (exceto as colunas de datas)
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            if idx < 9 or idx > 10:  # Exclui as colunas 9 e 10 (as de datas)
                cell.alignment = Alignment(horizontal='left',
                                           vertical='center')  # Alinhamento horizontal e vertical central

    # Salvar a planilha no response
    wb.save(response)

    return response


def excel_detalhado(request):

    query = request.GET.get('query', '')
    data_inicial = request.GET.get('dataInicial')
    data_final = request.GET.get('dataFinal')
    unidade = request.GET.get('unidade')
    carga_horaria = request.GET.get('carga_horaria')


    # Converte as datas em objetos datetime, se fornecidas
    data_inicial = parse_date(data_inicial) if data_inicial else None
    data_final = parse_date(data_final) if data_final else None

    # Filtrar os dados com base nos parâmetros
    queryset = Ajuda_Custo.objects.all()
    if query:
        queryset = queryset.filter(Q(nome__icontains=query) | Q(matricula__icontains=query))

    if unidade:
        queryset = queryset.filter(unidade=unidade)

    if carga_horaria:
        queryset = queryset.filter(carga_horaria=carga_horaria)

    if data_inicial and data_final:
        queryset = queryset.filter(data__range=[data_inicial, data_final])
    elif data_inicial:
        queryset = queryset.filter(data__gte=data_inicial)
    elif data_final:
        queryset = queryset.filter(data__lte=data_final)

    # Processar os resultados em um dicionário
    dados = []
    for item in queryset:
        matricula = item.matricula
        unidade = item.unidade
        nome = item.nome
        cpf = ''  # Deixe a coluna de CPF em branco se não disponível
        data = item.data

        carga_horaria = item.carga_horaria

        dados.append([
            matricula,
            unidade,
            cpf,
            nome,
            data,
            carga_horaria
        ])

    # Criar uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Completo Ajuda Custo"

    # Adicionar o período de datas no topo da planilha
    ws.merge_cells('A1:F1')
    ws[
        'A1'] = f"Período: {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}" if data_inicial and data_final else 'Período: Todos os Dados'

    # Adicionar os cabeçalhos das colunas
    cabecalhos = ['Matrícula', 'Unidade', 'CPF', 'Nome', 'Data', 'Carga Horária']
    ws.append(cabecalhos)

    # Adicionar os dados na planilha
    for linha in dados:
        ws.append(linha)

    # Ajustar a largura das colunas
    for col in ws.iter_cols(min_row=2, max_col=ws.max_column):
        max_length = 0
        column_letter = col[0].column_letter  # Obtém a letra da coluna (A, B, C, etc.)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Ajustar o alinhamento central para todas as células
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Salvar a planilha no response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ajuda_custo_detalhado.xlsx'
    wb.save(response)

    return response


class AjudaCusto(LoginRequiredMixin, ListView):
    model = Ajuda_Custo
    template_name = "ajuda_custo.html"
    context_object_name = 'datas'
    paginate_by = 50  # Quantidade de registros por página

    def get_queryset(self):
        user = self.request.user

        # Verificação dos grupos de usuário
        if user.groups.filter(name__in=['Administrador', 'GerGesipe']).exists():
            # Acesso completo para Administradores e GerGesipe
            queryset = Ajuda_Custo.objects.all()
        elif user.groups.filter(name='Gerente').exists():
            # Acesso limitado à unidade do gestor
            try:
                unidade_gestor = user.cotaajudacusto_set.first().unidade
                queryset = Ajuda_Custo.objects.filter(unidade=unidade_gestor)
            except AttributeError:
                # Caso o gestor não tenha uma unidade atribuída
                queryset = Ajuda_Custo.objects.none()
        else:
            # Acesso limitado ao próprio usuário
            queryset = Ajuda_Custo.objects.filter(matricula=user.matricula)

        # Captura os valores de ano e mês do formulário de filtro
        ano_selecionado = self.request.GET.get('ano', timezone.now().year)
        mes_selecionado = self.request.GET.get('mes', timezone.now().month)

        # Certifique-se de converter para int
        ano_selecionado = int(ano_selecionado)
        mes_selecionado = int(mes_selecionado)

        # Filtrar por ano e mês
        queryset = queryset.filter(data__year=ano_selecionado, data__month=mes_selecionado)


        # Filtrar por data (mês atual)
        today = now().date()
        first_day_of_month = today.replace(day=1)
        last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        # Aplicar filtro de data
        data_inicial = first_day_of_month
        data_final = last_day_of_month
        queryset = queryset.filter(data__range=[data_inicial, data_final])

        return queryset.order_by('nome')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Paginação personalizada
        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        # Lógica para limitar a exibição das páginas
        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)

        # Dicionário com os nomes dos meses
        meses = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
            7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }

        # Lista de anos disponíveis (últimos 5 anos e o atual)
        current_year = now().year
        anos = list(range(current_year - 5, current_year + 1))

        # Captura os valores de ano e mês enviados pelo formulário de filtro
        ano_selecionado = self.request.GET.get('ano', timezone.now().year)
        mes_selecionado = self.request.GET.get('mes', timezone.now().month)

        # Certifique-se de converter para `int` para usá-los em consultas e lógica
        ano_selecionado = int(ano_selecionado)
        mes_selecionado = int(mes_selecionado)

        # Atualizar o contexto com os filtros
        context.update({
            'meses': meses,
            'anos': anos,
            'ano_selecionado': ano_selecionado,
            'mes_selecionado': mes_selecionado,
        })

        # Construir o restante do contexto com os filtros aplicados
        return build_context(self.request, context, ano_selecionado, mes_selecionado)


class RelatorioAjudaCusto(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Ajuda_Custo
    template_name = "relatorio_ajuda_custo.html"
    context_object_name = 'datas'
    paginate_by = 50  # Quantidade de registros por página

    def test_func(self):
        user = self.request.user
        grupos_permitidos = ['Administrador', 'GerGesipe', 'Gerente']
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)

    def get_queryset(self):
        query = self.request.GET.get('query', '')
        data_inicial = self.request.GET.get('dataInicial')
        data_final = self.request.GET.get('dataFinal')

        # Caso dataInicial ou dataFinal não sejam fornecidas, utilizar o mês corrente
        today = now().date()
        first_day_of_month = today.replace(day=1)
        last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        # Garantir que as datas estejam no formato de string antes de usar parse_date
        if not data_inicial:
            data_inicial = first_day_of_month
        else:
            data_inicial = parse_date(data_inicial)  # Converter para objeto de data

        if not data_final:
            data_final = last_day_of_month
        else:
            data_final = parse_date(data_final)  # Converter para objeto de data

        queryset = Ajuda_Custo.objects.all()

        # Filtro por unidade do gerente (se o usuário for um gerente)
        user = self.request.user
        if user.groups.filter(name='Gerente').exists():
            # Obtém a unidade gerenciada pelo gestor
            cota_gerente = CotaAjudaCusto.objects.filter(gestor=user).first()
            if cota_gerente:
                queryset = queryset.filter(unidade=cota_gerente.unidade)

        if query:
            queryset = queryset.filter(
                Q(nome__icontains=query) | Q(matricula__icontains=query)
            )

        # Filtro por unidade (se fornecido)
        unidade = self.request.GET.get('unidade')
        if unidade:
            queryset = queryset.filter(unidade=unidade)  # Filtra pelo nome da unidade

        # Filtro por carga horária
        carga_horaria = self.request.GET.get('carga_horaria')
        if carga_horaria:
            queryset = queryset.filter(carga_horaria=carga_horaria)

        # Filtro por data
        if data_inicial and data_final:
            queryset = queryset.filter(data__range=[data_inicial, data_final])
        elif data_inicial:
            queryset = queryset.filter(data__gte=data_inicial)
        elif data_final:
            queryset = queryset.filter(data__lte=data_final)

        return queryset.order_by('nome')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Cria o formulário de filtros
        form = FiltroRelatorioForm(self.request.GET or None)

        # Verifica se o usuário é um gerente
        if self.request.user.groups.filter(name='Gerente').exists():
            # Obtém a unidade gerenciada pelo gestor
            cota_gerente = CotaAjudaCusto.objects.filter(gestor=self.request.user).first()
            if cota_gerente:
                # Define o valor inicial do campo unidade como a unidade do gerente
                form.fields['unidade'].initial = cota_gerente.unidade
                # Desabilita o campo unidade
                form.fields['unidade'].disabled = True

        # Adiciona o formulário ao contexto
        context['form'] = form

        # Obtém os nomes dos grupos do usuário
        context['user_groups'] = self.request.user.groups.values_list('name', flat=True)

        # Garantir que as datas estejam no formato correto (YYYY-MM-DD)
        today = now().date()
        first_day_of_month = today.replace(day=1)
        last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        data_inicial = self.request.GET.get('dataInicial', first_day_of_month)
        data_final = self.request.GET.get('dataFinal', last_day_of_month)

        # Certifique-se de que as datas estejam formatadas corretamente (YYYY-MM-DD)
        if isinstance(data_inicial, str):
            data_inicial = parse_date(data_inicial)  # Se for string, converter para data
        if isinstance(data_final, str):
            data_final = parse_date(data_final)  # Se for string, converter para data

        context['query'] = self.request.GET.get('query', '')
        context['dataInicial'] = data_inicial.strftime('%Y-%m-%d') if data_inicial else ''
        context['dataFinal'] = data_final.strftime('%Y-%m-%d') if data_final else ''
        context['unidades'] = Ajuda_Custo.objects.values_list('unidade', flat=True).distinct().order_by('nome')
        context['carga_horarias'] = Ajuda_Custo.objects.values_list('carga_horaria', flat=True).distinct()

        return context

    def get(self, request, *args, **kwargs):
        action = request.GET.get('action')
        queryset = self.get_queryset()

        if action == 'export_excel':
            return exportar_excel(request)
        elif action == 'excel_detalhado':
            return excel_detalhado(request)
        elif action == 'arquivos_assinados':
            return criar_arquivo_zip(request, queryset)
        elif action == 'gerar_pdf':
            return gerar_pdf_ajuda_custo(request, queryset)
        return super().get(request, *args, **kwargs)


class EnvioDatasView(LoginRequiredMixin, FormView):
    form_class = EnvioDatasForm
    template_name = 'ajuda_custo_add.html'
    success_url = reverse_lazy('ajuda_custo:confirmacao_datas')

    def form_valid(self, form):
        request = self.request

        mes = form.cleaned_data['mes']
        ano = form.cleaned_data['ano']
        dias = request.POST.getlist('dia')
        unidades = request.POST.getlist('unidade')
        cargas_horarias = request.POST.getlist('carga_horaria')


        try:
            servidor = get_servidor(request)
            print(f"Servidor encontrado: {servidor}")


            inicio_do_mes, fim_do_mes = get_intervalo_mes(int(mes), int(ano))
            print(f"Início do mês: {inicio_do_mes}, Fim do mês: {fim_do_mes}")

            registros_mes = get_registros_mes(servidor, inicio_do_mes, fim_do_mes)
            print(f"Registros do mês encontrados: {registros_mes.count()}")

            horas_por_unidade, horas_totais = calcular_horas_por_unidade(registros_mes)
            print("Horas por unidade no mês:", horas_por_unidade)
            print(f"Horas totais no mês: {horas_totais}")

            limites_horas_por_unidade = get_limites_horas_por_unidade(servidor)
            print("Limites de horas por unidade:", limites_horas_por_unidade)

            horas_a_adicionar_por_unidade = calcular_horas_a_adicionar_por_unidade(unidades, cargas_horarias)
            print("Horas a adicionar por unidade:", horas_a_adicionar_por_unidade)

            if not verificar_limites(horas_totais, horas_a_adicionar_por_unidade, horas_por_unidade, limites_horas_por_unidade, request):
                return self.form_invalid(form)

            # Salvar dados na sessão para a próxima view
            request.session['mes'] = mes
            request.session['ano'] = ano
            request.session['dias'] = dias
            request.session['unidades'] = unidades
            request.session['cargas_horarias'] = cargas_horarias

            # Gerar e enviar código de verificação
            codigo, expiracao = gerar_e_armazenar_codigo(request)
            print(f"Código gerado: {codigo}, expira em: {expiracao}")

            return redirect(self.success_url)

        except Exception as e:
            print(f"Erro interno: {e}")
            messages.error(request, 'Ocorreu um erro interno. Tente novamente mais tarde.')
            return self.form_invalid(form)


@login_required
def reenviar_codigo(request):
    if request.method == "POST":
        try:
            # Gerar e enviar novo código usando a função centralizada
            novo_codigo, expiracao = gerar_e_armazenar_codigo(request)
            print(f"Novo código de verificação gerado: {novo_codigo}, expira em: {expiracao}")

            return JsonResponse({
                "success": True,
                "message": "Novo código de verificação enviado com sucesso.",
                "expiracao": expiracao.isoformat()  # Retornar o tempo de expiração
            }, status=200)
        except Exception as e:
            print(f"Erro ao reenviar código: {e}")
            return JsonResponse({"error": "Erro ao reenviar o código. Tente novamente mais tarde."}, status=500)

    return JsonResponse({"error": "Método não permitido"}, status=405)


class ConfirmacaoDatasView(LoginRequiredMixin, FormView):
    form_class = ConfirmacaoDatasForm
    template_name = 'confirmacao_datas.html'
    success_url = reverse_lazy('ajuda_custo:ajuda_custo')

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        print('codigo_verificacao')

        # Lógica para reenviar código
        if 'reenviar_codigo' in request.POST:
            try:
                response = reenviar_codigo(request)  # Chama a função reenviar_codigo
                if response.status_code == 200:
                    messages.info(request, 'Um novo código de verificação foi enviado para seu e-mail.')
                else:
                    messages.error(request, 'Ocorreu um erro ao tentar reenviar o código.')
            except Exception as e:
                print(f"Erro ao reenviar código: {e}")
                messages.error(request, 'Ocorreu um erro ao tentar reenviar o código.')

            return self.form_invalid(form)  # Mantém o formulário aberto para inserção do novo código

        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        request = self.request

        codigo_verificacao = form.cleaned_data['codigo_verificacao']
        print(f"Código de verificação recebido: {codigo_verificacao}")

        if not codigo_verificacao:
            messages.error(request, 'Por favor, insira o código de verificação.')
            print("Erro: Código de verificação não fornecido.")
            return self.form_invalid(form)

        expira_em = request.session.get('codigo_verificacao_expira', '')
        print(f"Código expira em: {expira_em}")

        if not verificar_codigo_expiracao(expira_em):
            messages.error(request, 'O código de verificação expirou. Envie um novo código.')
            print("Erro: Código de verificação expirou.")
            return self.form_invalid(form)

        if codigo_verificacao != request.session.get('codigo_verificacao'):
            messages.error(request, 'Código de verificação incorreto. Envie um novo código.')
            print("Erro: Código de verificação incorreto.")
            return self.form_invalid(form)

        print("Código de verificação válido. Processando formulário...")

        try:
            mes = request.session['mes']
            ano = request.session['ano']
            dias = request.session['dias']
            unidades = request.session['unidades']
            cargas_horarias = request.session['cargas_horarias']
            servidor = get_servidor(request)
            print(f'os dias {dias,} os meses {mes}, as cargas horarias {cargas_horarias}, as unidades {unidades}')

            print(servidor)
            nome_servidor = servidor.nome

            # Processa as datas e salva os registros
            ajuda_custo_instances, error_messages = self.processar_datas(dias, unidades, cargas_horarias, servidor,
                                                                         nome_servidor, mes, ano, request)
            self.exibir_mensagens(error_messages)

            # Salva o código de verificação em cada instância de Ajuda_Custo
            for ajuda_custo in ajuda_custo_instances:
                ajuda_custo.codigo_verificacao = codigo_verificacao
                ajuda_custo.save()

            self.enviar_email(ajuda_custo_instances, error_messages)

            return redirect(self.success_url)

        except Exception as e:
            print(f"Erro interno: {e}")
            messages.error(request, 'Ocorreu um erro interno. Tente novamente mais tarde.')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        mes = request.session.get('mes')
        ano = request.session.get('ano')
        dias = request.session.get('dias', [])
        unidades = request.session.get('unidades', [])
        cargas_horarias = request.session.get('cargas_horarias', [])
        servidor = get_object_or_404(Servidor, matricula=request.user.matricula)
        nome_servidor = servidor.nome
        matricula = servidor.matricula

        print(request.session.keys())
        print(dias)
        print(request.session.get('unidades'))
        print(cargas_horarias)
        print(mes)
        print(ano)
        print(nome_servidor)
        print(servidor)

        context['matricula'] = matricula

        context['datas'] = zip(dias, unidades, cargas_horarias) # Criar lista de tuplas return

        ajuda_custo_instances, error_messages = datas_adicionar(
            dias, unidades, cargas_horarias, servidor, nome_servidor, mes, ano, request
        )

        context['ajuda_custo_instances'] = ajuda_custo_instances
        context['error_messages'] = error_messages
        print(ajuda_custo_instances)

        return context

    def gerar_novo_codigo(self):
        request = self.request
        # Gerar e enviar código de verificação
        codigo, expiracao = gerar_e_armazenar_codigo(request)
        print(f"Código gerado: {codigo}, expira em: {expiracao}")

    @staticmethod
    def processar_datas(dias, unidades, cargas_horarias, servidor, nome_servidor, mes, ano, request):
        error_messages = []
        ajuda_custo_instances = []

        with transaction.atomic():
            for dia, unidade_nome, carga_horaria in zip(dias, unidades, cargas_horarias):
                try:
                    data_completa = datetime.strptime(f"{dia}/{mes}/{ano}", "%d/%m/%Y").date()
                    print(f"Processando data: {data_completa}")
                    horas_a_adicionar = int(carga_horaria.strip().replace(' horas', ''))
                    print(f"Horas a adicionar: {horas_a_adicionar}")

                    if Ajuda_Custo.objects.filter(matricula=servidor.matricula, data=data_completa).exists():
                        error_messages.append(f'O servidor já possui uma entrada para {dia.strip()}/{mes}/{ano}.')
                        print(f"Data duplicada detectada: {data_completa}")
                        continue

                    carga_horaria_final = f"{horas_a_adicionar} horas"
                    ajuda_custo = Ajuda_Custo(
                        matricula=servidor.matricula,
                        nome=nome_servidor,
                        data=data_completa,
                        unidade=unidade_nome.strip(),
                        carga_horaria=carga_horaria_final,
                        majorado=DataMajorada.objects.filter(data=data_completa).exists()
                    )
                    ajuda_custo.save()
                    ajuda_custo_instances.append(ajuda_custo)
                    print(f"Registro salvo: {ajuda_custo}")

                except Exception as e:
                    print(f"Erro ao processar data {dia.strip()}/{mes}/{ano}: {e}")
                    error_messages.append(f'Ocorreu um erro ao processar a data {dia.strip()}/{mes}/{ano}.')

        return ajuda_custo_instances, error_messages

    def exibir_mensagens(self, error_messages):
        for message in error_messages:
            messages.error(self.request, message)
        if not error_messages:
            messages.success(self.request, 'Datas adicionadas com sucesso!')
        else:
            messages.warning(self.request, 'Processamento concluído com algumas falhas. Verifique as mensagens de erro.')

    def enviar_email(self, ajuda_custo_instances, error_messages):
        email_destinatario = self.request.user.email
        print(f"Enviando email para {email_destinatario}")

        assunto = 'Confirmação de Datas Marcadas'
        contexto = {'ajuda_custo_list': ajuda_custo_instances, 'servidor': self.request.user.nome_completo, 'matricula': self.request.user.matricula, 'error_messages': error_messages}
        corpo_email = render_to_string('email_datasmarcadas.html', contexto)
        email = EmailMessage(assunto, corpo_email, settings.EMAIL_HOST_USER, [email_destinatario])
        email.content_subtype = 'html'
        email.send()
        print('Email enviado com sucesso')


class AdminCadastrar(LoginRequiredMixin, UserPassesTestMixin, FormView):
    model = Ajuda_Custo
    form_class = AdminDatasForm
    template_name = 'admin_cadastrar.html'
    success_url = reverse_lazy('ajuda_custo:admin_cadastrar')

    def test_func(self):
        user = self.request.user
        # Define os grupos permitidos
        grupos_permitidos = ['Administrador', 'GerGesipe']
        # Retorna True se o usuário pertence a pelo menos um dos grupos
        return user.groups.filter(name__in=grupos_permitidos).exists()

        # Levanta exceção em caso de falta de permissão

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)  # Substitua '404.html' pelo nome do seu template

    def form_valid(self, form):
        mes = form.cleaned_data['mes']
        ano = form.cleaned_data['ano']
        unidade = form.cleaned_data['unidade']
        dias_12h = form.cleaned_data['dias_12h']
        dias_24h = form.cleaned_data['dias_24h']

        dias_12h_list = [dia.strip() for dia in dias_12h.split(',') if dia.strip()]
        dias_24h_list = [dia.strip() for dia in dias_24h.split(',') if dia.strip()]

        # Pegando a matrícula do campo do formulário (presumindo que o campo seja 'matricula')
        matricula = self.request.POST.get('matricula')

        # Obtém o servidor com base na matrícula fornecida no formulário
        try:
            servidor = Servidor.objects.get(matricula=matricula)
        except Servidor.DoesNotExist:
            messages.error(self.request, 'Erro: Servidor não encontrado.')
            return redirect(self.success_url)

        mes_int = int(mes)
        ano_int = int(ano)

        # Verifica o total de horas mensais antes de adicionar cada registro
        for dia in dias_12h_list + dias_24h_list:
            try:
                data_completa = datetime.strptime(f"{dia}/{mes}/{ano}", "%d/%m/%Y").date()

                # Verifique se o servidor já marcou essa data, independentemente da carga horária
                if Ajuda_Custo.objects.filter(matricula=servidor.matricula, data=data_completa).exists():
                    messages.error(self.request, f'O servidor já possui uma entrada para {dia}/{mes}/{ano}.')
                    return redirect(self.success_url)

                inicio_do_mes = data_completa.replace(day=1)
                fim_do_mes = (inicio_do_mes + timedelta(days=31)).replace(day=1) - timedelta(days=1)

                registros_mes = Ajuda_Custo.objects.filter(
                    matricula=servidor.matricula,
                    data__range=[inicio_do_mes, fim_do_mes]
                )

                total_horas_mes = 0
                for registro in registros_mes:
                    if registro.carga_horaria == '12 horas':
                        total_horas_mes += 12
                    elif registro.carga_horaria == '24 horas':
                        total_horas_mes += 24

                horas_a_adicionar = 12 if dia in dias_12h_list else 24

                if total_horas_mes + horas_a_adicionar > 192:
                    messages.error(self.request, f'Limite de 192 horas mensais excedido para {dia}/{mes}/{ano}.')
                    return redirect(self.success_url)

                majorado = DataMajorada.objects.filter(data=data_completa).exists()

                ajuda_custo = Ajuda_Custo(
                    matricula=servidor.matricula,
                    nome=servidor.nome,
                    data=data_completa,
                    unidade=unidade,
                    carga_horaria='12 horas' if dia in dias_12h_list else '24 horas',
                    majorado=majorado
                )
                ajuda_custo.save()

            except ValueError:
                messages.error(self.request, f'Erro: Data inválida - {dia}/{mes}/{ano}')
                return redirect(self.success_url)

        messages.success(self.request, 'Datas adicionadas com sucesso!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Erro no Cadastro, Confira os Dados e Tente Novamente.')
        return super().form_invalid(form)


class HorasLimite(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'horas_limite.html'
    success_url = reverse_lazy('ajuda_custo:horas_limite')

    def test_func(self):
        user = self.request.user
        grupos_permitidos = ['Administrador', 'GerGesipe']
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)

    def get_form_class(self):
        if self.request.method == 'POST':
            if 'servidor' in self.request.POST:
                return LimiteAjudaCustoForm
            else:
                return CotaAjudaCustoForm
        return LimiteAjudaCustoForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        query = self.request.GET.get('query', '')
        unidade = self.request.GET.get('unidade', '')

        # Paginação para LimiteAjudaCusto (aba Individual)
        filtros_cargas = Q()
        if query:
            filtros_cargas &= Q(servidor__nome__icontains=query) | Q(servidor__matricula__icontains=query)
        if unidade:
            filtros_cargas &= Q(unidade__nome=unidade)

        cargas = LimiteAjudaCusto.objects.filter(filtros_cargas).order_by('servidor')

        paginator_cargas = Paginator(cargas, 20)
        page_number_cargas = self.request.GET.get('page_cargas')
        page_obj_cargas = paginator_cargas.get_page(page_number_cargas)

        # Paginação para CotaAjudaCusto (aba Gerências)

        cotas = CotaAjudaCusto.objects.all().order_by('unidade')
        paginator_cotas = Paginator(cotas, 100)
        page_number_cotas = self.request.GET.get('page_cotas')
        page_obj_cotas = paginator_cotas.get_page(page_number_cotas)

        # Adiciona os dados ao contexto
        context.update({
            'unidades': Unidade.objects.values_list('nome', flat=True).order_by('nome'),
            'carga_horaria': page_obj_cargas,
            'cota_horaria': page_obj_cotas,
            'page_obj_cargas': page_obj_cargas,
            'page_obj_cotas': page_obj_cotas,
            'query': query,
            'unidade': unidade,
        })

        # Adiciona os formulários para carga horária individual e por gerências
        context['form_individual'] = LimiteAjudaCustoForm()
        context['form_gerencia'] = CotaAjudaCustoForm()

        return context

    def post(self, request, *args, **kwargs):
        if 'servidor' in request.POST:
            form = LimiteAjudaCustoForm(request.POST)
            if form.is_valid():
                return self.form_individual_valid(form)
            else:
                return self.form_invalid(form)
        else:
            form = CotaAjudaCustoForm(request.POST)
            if form.is_valid():
                return self.form_gerencia_valid(form)
            else:
                return self.form_invalid(form)

    def form_individual_valid(self, form):
        servidor = form.cleaned_data['servidor']
        unidade = form.cleaned_data['unidade']
        limite_horas = form.cleaned_data['limite_horas']

        # Verifica se a carga horária excede 192 horas
        if limite_horas > 192:
            messages.error(self.request, "A carga horária não pode ultrapassar 192 horas.")
            return self.form_invalid(form)

        LimiteAjudaCusto.objects.update_or_create(
            servidor=servidor,
            unidade=unidade,
            defaults={'limite_horas': limite_horas}
        )

        messages.success(self.request, 'Limite de horas atualizado com sucesso!')
        return redirect(self.success_url)

    def form_gerencia_valid(self, form):
        gestor = form.cleaned_data['gestor']
        unidade = form.cleaned_data['unidade']
        cota_ajudacusto = form.cleaned_data['cota_ajudacusto']

        CotaAjudaCusto.objects.update_or_create(
            gestor=gestor,
            unidade=unidade,
            defaults={'cota_ajudacusto': cota_ajudacusto}
        )

        messages.success(self.request, 'Cota de ajuda de custo atualizada com sucesso!')
        return redirect(self.success_url)


class CargaHorariaGerente(LoginRequiredMixin, UserPassesTestMixin, FormView):
    model = LimiteAjudaCusto
    form_class = LimiteAjudaCustoForm
    template_name = 'cargahoraria_gerente.html'
    success_url = reverse_lazy('ajuda_custo:cargahoraria_gerente')

    # Verifica se o usuário pertence a determinados grupos
    def test_func(self):
        user = self.request.user
        # Define os grupos permitidos
        grupos_permitidos = ['Administrador', 'Gerente']
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)

    def form_valid(self, form):
        gestor = self.request.user
        unidade = form.cleaned_data['unidade']
        limite_horas = form.cleaned_data['limite_horas']
        servidor = form.cleaned_data['servidor']

        # Obtém a matrícula do gestor (usuário logado)
        matricula_gestor = gestor.matricula

        # Verifica se o servidor selecionado é o mesmo que o servidor do gestor
        if str(servidor.matricula).strip() == str(matricula_gestor).strip():
            messages.error(self.request, "Você não pode atribuir cota de horas para si mesmo.")
            return self.form_invalid(form)

        # Verifica se a carga horária excede 192 horas
        if limite_horas > 192:
            messages.error(self.request, "A carga horária não pode ultrapassar 192 horas.")
            return self.form_invalid(form)

        # Verifica a cota disponível do gerente
        cota = CotaAjudaCusto.objects.get(gestor=gestor, unidade=unidade)
        if limite_horas > cota.carga_horaria_disponivel:
            messages.error(self.request, f"Você não pode distribuir mais que {cota.carga_horaria_disponivel} horas.")
            return self.form_invalid(form)

        # Atualiza ou cria o limite de horas
        LimiteAjudaCusto.objects.update_or_create(
            servidor=servidor,
            unidade=unidade,
            defaults={'limite_horas': limite_horas}
        )

        # Atualiza a cota disponível
        cota.save()

        messages.success(self.request, 'Carga horária distribuída com sucesso!')
        return redirect(self.success_url)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Adiciona a unidade baseada no gestor
        kwargs['initial'] = {'unidade': self.request.user.cotaajudacusto_set.first().unidade}
        return kwargs

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Torna o campo unidade somente leitura
        form.fields['unidade'].disabled = True
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtém o valor da consulta de pesquisa, se houver
        query = self.request.GET.get('query', '')
        unidade_nome = self.request.GET.get('unidade', '')
        context['cota_total'] = self.request.user.cotaajudacusto_set.first().carga_horaria_total
        context['cota_disponivel'] = self.request.user.cotaajudacusto_set.first().carga_horaria_disponivel

        # Filtra por nome, matrícula e unidade, se houver uma consulta
        filtros = Q()
        if query:
            filtros &= Q(servidor__nome__icontains=query) | Q(servidor__matricula__icontains=query)

        # Valida o nome da unidade antes de aplicar na query
        if unidade_nome:
            try:
                unidade = Unidade.objects.get(nome=unidade_nome)
                filtros &= Q(unidade=unidade)
            except Unidade.DoesNotExist:
                messages.error(self.request, f"Unidade '{unidade_nome}' não encontrada.")
                filtros &= Q(pk__isnull=True)  # Retorna nenhum resultado

        # Filtra as cargas horárias com base no gestor logado
        cargas = LimiteAjudaCusto.objects.filter(filtros, unidade__cotaajudacusto__gestor=self.request.user).order_by('servidor')

        # Limita a carga horária máxima a 192 horas
        cargas = cargas.filter(limite_horas__lte=192)

        # Adiciona a paginação
        paginator = Paginator(cargas, 20)  # Mostra 20 registros por página
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Adiciona os dados ao contexto
        context['unidades'] = Unidade.objects.values_list('nome', flat=True).order_by('nome')
        context['carga_horaria'] = page_obj
        context['page_obj'] = page_obj
        context['query'] = query
        context['unidade'] = unidade_nome

        return context


#logicas de exclusao
def excluir_ajuda_custo(request, pk):
    ajuda_custo = get_object_or_404(Ajuda_Custo, pk=pk)
    user = request.user

    # Verifica se o usuário é Administrador ou GerGesipe (sem restrições)
    if user.groups.filter(name__in=['Administrador', 'GerGesipe']).exists():
        ajuda_custo.delete()
        messages.success(request, 'Ajuda de custo excluída com sucesso!')
        return redirect('ajuda_custo:relatorio_ajuda_custo')

    # Verifica se o usuário é Gerente
    elif user.groups.filter(name='Gerente').exists():
        # Obtém a data atual
        hoje = timezone.now().date()
        data_ajuda = ajuda_custo.data

        # Verifica se a ajuda de custo é do mês atual ou de um mês futuro
        if (data_ajuda.year > hoje.year) or (data_ajuda.year == hoje.year and data_ajuda.month >= hoje.month):
            ajuda_custo.delete()
            messages.success(request, 'Ajuda de custo excluída com sucesso!')
        else:
            messages.error(request,
                           'Você só pode excluir ajudas de custo do mês atual ou futuros. Para exclusões anteriores, entre em contato com o administrador.')

        return redirect('ajuda_custo:relatorio_ajuda_custo')

    # Caso o usuário não tenha permissão
    else:
        messages.error(request, 'Você não tem permissão para excluir ajudas de custo.')
        return redirect('ajuda_custo:relatorio_ajuda_custo')



def excluir_cota(request, pk):
    cota = get_object_or_404(CotaAjudaCusto, pk=pk)
    cota.delete()
    messages.success(request, 'Cota de horas excluído com sucesso!')
    return redirect('ajuda_custo:horas_limite')  # Redirecionamento padrão


def excluir_limite(request, pk):
    limite = get_object_or_404(LimiteAjudaCusto, pk=pk)
    unidade = limite.unidade
    limite_horas = limite.limite_horas
    limite.delete()

    # Atualize a carga horária disponível da gerência correspondente
    cota = CotaAjudaCusto.objects.filter(unidade=unidade).first()
    if cota:
        cota.carga_horaria_disponivel += limite_horas
        cota.save()

    messages.success(request, 'Limite de horas excluído com sucesso!')

    # Redireciona com base no grupo do usuário
    if request.user.groups.filter(name='Gerente').exists():
        return redirect('ajuda_custo:cargahoraria_gerente')
    elif request.user.groups.filter(name='Administrador').exists():
        return redirect('ajuda_custo:horas_limite')
    else:
        return redirect('ajuda_custo:horas_limite')  # Redirecionamento padrão


class VerificarCargaHoraria(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Ajuda_Custo
    template_name = "verificar_carga_horaria.html"
    context_object_name = "dados"
    paginate_by = 50

    def test_func(self):
        user = self.request.user
        grupos_permitidos = ["Administrador", "GerGesipe"]
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, "403.html", status=403)

    def get_queryset(self):
        mes = self.request.GET.get("mes")
        ano = self.request.GET.get("ano")
        query = self.request.GET.get("query", "").strip()

        queryset = Ajuda_Custo.objects.all()

        if mes and ano:
            queryset = queryset.filter(data__year=ano, data__month=mes)

        if query:
            queryset = queryset.filter(nome__icontains=query)

        return queryset.order_by("nome")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes = self.request.GET.get("mes")
        ano = self.request.GET.get("ano")
        query = self.request.GET.get("query", "").strip()
        today = now().date()

        if not mes:
            mes = today.month
        if not ano:
            ano = today.year

        # 1. Obter matrículas importantes (com filtro de query)
        matriculas_importantes = MatriculaImportante.objects.select_related('servidor').all()
        if query:
            matriculas_importantes = matriculas_importantes.filter(
                Q(matricula__icontains=query) |
                Q(servidor__nome__icontains=query)
            )
        matriculas_importantes_lista = matriculas_importantes.order_by('servidor__nome')
        matriculas_importantes_values = matriculas_importantes.values_list('matricula', flat=True)

        # 2. Obter servidores presentes no mês (com filtro de query)
        servidores_presentes = Ajuda_Custo.objects.filter(
            data__year=ano,
            data__month=mes
        )
        if query:
            servidores_presentes = servidores_presentes.filter(nome__icontains=query)
        servidores_presentes_values = servidores_presentes.values_list('matricula', flat=True).distinct()

        # 3. Identificar matrículas importantes faltantes (convertendo para string para comparação)
        matriculas_importantes_str = {str(m) for m in matriculas_importantes_values}
        servidores_presentes_str = {str(s) for s in servidores_presentes_values}
        matriculas_faltantes = matriculas_importantes_str - servidores_presentes_str

        # 4. Obter dados agregados por servidor (total de horas)
        dados_agregados = (
            Ajuda_Custo.objects.filter(data__year=ano, data__month=mes)
            .values("matricula", "nome")
            .annotate(
                total_horas=Sum(
                    Case(
                        When(carga_horaria="12 horas", then=12),
                        When(carga_horaria="24 horas", then=24),
                        default=0,
                        output_field=IntegerField(),
                    )
                )
            )
        )

        # Criar um dicionário rápido para consulta de horas por matrícula
        horas_por_matricula = {str(item['matricula']): item['total_horas'] for item in dados_agregados}

        # 5. Obter dados dos faltantes
        faltantes = []
        for matricula in matriculas_faltantes:
            try:
                info = matriculas_importantes.get(matricula=matricula)
                # Usar as horas do dicionário ou 0 se não existir
                total_horas = horas_por_matricula.get(matricula, 0)

                faltantes.append({
                    'matricula': matricula,
                    'nome': info.servidor.nome if info.servidor else "Nome não encontrado",
                    'erro': 'Servidor importante faltante',
                    'total_horas': total_horas,  # Agora mostra o total real de horas
                    'limite_horas': 192,
                    'datas_repetidas': []
                })
            except MatriculaImportante.DoesNotExist:
                pass

        # 6. Verificar datas repetidas para cada servidor
        servidores_com_datas_repetidas = (
            Ajuda_Custo.objects.filter(data__year=ano, data__month=mes)
            .values('matricula', 'nome', 'data')
            .annotate(total=Count('id'))
            .filter(total__gt=1)
            .order_by('matricula', 'data')
        )

        # Processar problemas
        servidores_problema = {}

        # Adicionar faltantes
        for faltante in faltantes:
            servidores_problema[faltante['matricula']] = faltante

        # Processar servidores que ultrapassaram 192 horas
        for servidor in dados_agregados:
            matricula = str(servidor["matricula"])
            if servidor["total_horas"] > 192:
                if matricula in servidores_problema:
                    servidores_problema[matricula]['erro'] += ' e limite excedido'
                    servidores_problema[matricula]['total_horas'] = servidor["total_horas"]
                else:
                    servidores_problema[matricula] = {
                        'matricula': matricula,
                        'nome': servidor["nome"],
                        'total_horas': servidor["total_horas"],
                        'limite_horas': 192,
                        'erro': 'Limite excedido',
                        'datas_repetidas': []
                    }

        # Processar servidores com datas repetidas
        for repeticao in servidores_com_datas_repetidas:
            matricula = str(repeticao["matricula"])
            data_repetida = repeticao["data"].strftime('%d/%m/%Y')

            if matricula in servidores_problema:
                if 'datas_repetidas' in servidores_problema[matricula]:
                    servidores_problema[matricula]['datas_repetidas'].append(data_repetida)
                else:
                    servidores_problema[matricula]['datas_repetidas'] = [data_repetida]

                if 'Limite excedido' in servidores_problema[matricula]['erro']:
                    servidores_problema[matricula]['erro'] = 'Limite excedido e datas repetidas'
                else:
                    servidores_problema[matricula]['erro'] = 'Datas repetidas'
            else:
                # Obter o total de horas para este servidor
                total_horas = horas_por_matricula.get(matricula, 0)

                servidores_problema[matricula] = {
                    'matricula': matricula,
                    'nome': repeticao["nome"],
                    'total_horas': total_horas,
                    'limite_horas': 192,
                    'erro': 'Datas repetidas',
                    'datas_repetidas': [data_repetida]
                }

        # Converter para lista e ordenar
        dados_com_problemas = []
        for matricula, info in servidores_problema.items():
            dados_com_problemas.append({
                'matricula': matricula,
                'nome': info['nome'],
                'total_horas': info.get('total_horas', 0),
                'limite_horas': info.get('limite_horas', 192),
                'erro': info['erro'],
                'datas_repetidas': ', '.join(info.get('datas_repetidas', [])) if info.get('datas_repetidas') else 'N/A'
            })

        # Ordenar por tipo de erro e nome
        ordem_erro = {
            'Servidor importante faltante e limite excedido e datas repetidas': 0,
            'Servidor importante faltante e limite excedido': 1,
            'Servidor importante faltante e datas repetidas': 2,
            'Servidor importante faltante': 3,
            'Limite excedido e datas repetidas': 4,
            'Limite excedido': 5,
            'Datas repetidas': 6
        }

        dados_com_problemas.sort(key=lambda x: (
            ordem_erro.get(x['erro'], 7),
            x['nome']
        ))

        # Configuração do contexto
        meses_nomes = {
            1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio",
            6: "Junho", 7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro",
            11: "Novembro", 12: "Dezembro"
        }
        # Adicione esta parte para a paginação
        paginator = Paginator(dados_com_problemas, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context.update({
            'dados': page_obj,
            'mes': int(mes),
            'ano': int(ano),
            'query': query,
            'meses': [(num, nome) for num, nome in meses_nomes.items()],
            'anos': range(today.year - 5, today.year + 1),
            'mes_atual_nome': meses_nomes[int(mes)],
            'total_erros': len(dados_com_problemas),
            'total_faltantes': len(faltantes),
            'matriculas_importantes': matriculas_importantes_lista
        })

        return context


def adicionar_matricula_importante(request):
    if request.method == 'POST':
        matricula = request.POST.get('matricula')

        # Busca o servidor pela matrícula (assumindo que a matrícula existe na tabela Servidor)
        try:
            servidor = Servidor.objects.get(matricula=matricula)

            # Cria a matrícula importante relacionada ao servidor
            MatriculaImportante.objects.create(
                matricula=matricula,
                servidor=servidor
            )
            messages.success(request, "Matrícula importante adicionada com sucesso!")

        except Servidor.DoesNotExist:
            messages.error(request, "Servidor com esta matrícula não encontrado!")
        except Exception as e:
            messages.error(request, f"Erro ao adicionar matrícula: {str(e)}")

    return redirect('ajuda_custo:verificar_carga_horaria')


def remover_matricula_importante(request, matricula_id):
    try:
        matricula = MatriculaImportante.objects.get(id=matricula_id)
        matricula.delete()
        messages.success(request, "Matrícula importante removida com sucesso!")
    except MatriculaImportante.DoesNotExist:
        messages.error(request, "Matrícula não encontrada!")

    return redirect('ajuda_custo:verificar_carga_horaria')