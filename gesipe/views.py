from django.shortcuts import render, redirect
from .models import Gesipe_adm, Gesipe_Sga
from django.views.generic.edit import FormView, UpdateView, CreateView, View
from django.utils.timezone import now
from .forms import GesipeAdmForm, UploadFileForm, GesipeSgaForm
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F  # Para pesquisas com OR lógico
from django.utils import timezone
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.http import JsonResponse
from datetime import datetime, date
import openpyxl
from servidor.models import Servidor
from io import BytesIO
import zipfile
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Image, Frame
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.staticfiles import finders
from django.contrib.auth.mixins import UserPassesTestMixin
from django.template.loader import get_template
from weasyprint import HTML



# Create your views here.

# url - view - html


#exportar relatorios para pdf

def export_pdf(request):
    # Pega as datas de filtro do request recebido
    data_inicial = request.GET.get('dataInicial')
    data_final = request.GET.get('dataFinal')

    # Filtra o queryset com as datas fornecidas
    queryset = RelatorioGesipeAdm().get_queryset()  # Chama o queryset da view manualmente, se necessário
    if data_inicial and data_final:
        try:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
            data_final = datetime.strptime(data_final, '%Y-%m-%d').date()
            queryset = queryset.filter(data__range=(data_inicial, data_final))
        except ValueError:
            pass  # Ignora erros de formato de data

    # Renderiza o template HTML com os dados filtrados
    template = get_template('relatorio_gesipe_pdf.html')
    context = {
        'object_list': queryset,  # Dados filtrados para o PDF
        'dataInicial': data_inicial,
        'dataFinal': data_final,
        'setor': 'Administrativo',
    }
    html = template.render(context)

    # Gera e retorna o PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename='relatorio_gesipe_adm.pdf'"
    HTML(string=html).write_pdf(response)
    return response


def export_pdf_sga_detalhado(request, setor):
    # Pega as datas de filtro do request recebido
    data_inicial = request.GET.get('dataInicial')
    data_final = request.GET.get('dataFinal')

    # Filtra o queryset com as datas fornecidas
    queryset = RelatorioGesipeSga().get_queryset()  # Chama o queryset da view manualmente, se necessário
    if data_inicial and data_final:
        try:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
            data_final = datetime.strptime(data_final, '%Y-%m-%d').date()
            queryset = queryset.filter(data__range=(data_inicial, data_final))
        except ValueError:
            pass  # Ignora erros de formato de data

    # Renderiza o template HTML com os dados filtrados
    template = get_template('sga/relatorio_gesipe_sga_pdf_detalhado.html')
    context = {
        'object_list': queryset,  # Dados filtrados para o PDF
        'dataInicial': data_inicial,
        'dataFinal': data_final,
        'setor': setor,
    }
    html = template.render(context)

    # Gera e retorna o PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename='relatorio_gesipe_sga_detalhado.pdf'"
    HTML(string=html).write_pdf(response)
    return response



class GesipeArmaria(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Servidor
    template_name = "gesipe_armaria.html"
    context_object_name = 'servidores'
    paginate_by = 20
    max_pdfs = 50  # Limite máximo de PDFs

    def test_func(self):
        user = self.request.user
        grupos_permitidos = ['Administrador', 'Armaria']
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)

    def get_queryset(self):
        queryset = Servidor.objects.all()
        query = self.request.GET.get('query')
        if query:
            queryset = queryset.filter(
                Q(nome__icontains=query) | Q(matricula__icontains=query)
            )

        cargo = self.request.GET.get('cargo')
        if cargo:
            queryset = queryset.filter(cargo=cargo)

        local_trabalho = self.request.GET.get('local_trabalho')
        if local_trabalho:
            queryset = queryset.filter(local_trabalho__icontains=local_trabalho)

        cargo_comissionado = self.request.GET.get('cargo_comissionado')
        if cargo_comissionado == "None":
            cargo_comissionado = ""

        if cargo_comissionado:
            queryset = queryset.filter(cargo_comissionado__icontains=cargo_comissionado)

        # Novo filtro de intervalo alfabético
        intervalo_alfabetico = self.request.GET.get('intervalo_alfabetico')
        if intervalo_alfabetico:
            if intervalo_alfabetico == 'A-E':
                queryset = queryset.filter(nome__regex=r'^[A-Ea-e]')
            elif intervalo_alfabetico == 'F-J':
                queryset = queryset.filter(nome__regex=r'^[F-Jf-j]')
            elif intervalo_alfabetico == 'K-O':
                queryset = queryset.filter(nome__regex=r'^[K-Ok-o]')
            elif intervalo_alfabetico == 'P-T':
                queryset = queryset.filter(nome__regex=r'^[P-Tp-t]')
            elif intervalo_alfabetico == 'U-Z':
                queryset = queryset.filter(nome__regex=r'^[U-Zu-z]')

        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        genero = self.request.GET.get('genero')
        if genero:
            queryset = queryset.filter(genero__icontains=genero)

        return queryset.order_by('nome')


    def generate_pdf(self, servidor):
        # Definindo a largura e altura da página
        page_width, page_height = A4

        # Crie o buffer e o canvas
        pdf_buffer = BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=(page_width, page_height))

        # Definindo margens
        margin_x = 2 * cm
        margin_y = page_height - 4 * cm  # Margem superior

        # Cabeçalho com as imagens
        logo_pb_path = finders.find('images/mini/GovPBT_mini.png')
        logo_seap_path = finders.find('images/mini/seap-pb_mini.png')
        logo_pp_path = finders.find('images/mini/pp_mini.png')

        def resize_image(img_path, max_width, max_height):
            img = Image(img_path)
            img.drawWidth = min(img.imageWidth, max_width)
            img.drawHeight = min(img.imageHeight, max_height)
            return img

        # Ajustar o tamanho das imagens
        logo_pb = resize_image(logo_pb_path, 2 * cm, 1 * cm)
        logo_seap = resize_image(logo_seap_path, 1 * cm, 1 * cm)
        logo_pp = resize_image(logo_pp_path, 2 * cm, 1 * cm)

        # Posicionamento das imagens com espaçamento
        start_x = margin_x
        start_y = margin_y - logo_pb.drawHeight
        spacing = 6 * cm

        logo_pb.drawOn(c, start_x, start_y)
        logo_seap.drawOn(c, start_x + logo_pb.drawWidth + spacing, start_y)
        logo_pp.drawOn(c, start_x + logo_pb.drawWidth + logo_seap.drawWidth + 2 * spacing, start_y)

        # Adicionando título centralizado
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(page_width / 2, start_y - 4 * cm, "TERMO DE ACAUTELAMENTO DE COLETE BALÍSTICO")

        # Conteúdo principal do texto com quebra automática de linha
        styles = getSampleStyleSheet()
        text_style = styles["BodyText"]
        text_style.leading = 22  # Ajuste o espaçamento entre linhas aqui
        text = (
            f"Eu, {servidor.nome}, Policial Penal do Estado da Paraíba, matrícula {servidor.matricula}, "
            f"lotado(a) atualmente no(a) {servidor.local_trabalho}, RECEBI da Gerência Executiva do Sistema "
            "Penitenciário, um (01) colete balístico da Marca PROTECTA, o qual ficará sob minha guarda e responsabilidade, "
            "ressarcindo o erário por sua perda ou extravio, independente dos procedimentos que venham a ocorrer no "
            "âmbito administrativo e criminal."
        )

        # Usando Frame e Paragraph para a quebra de linha automática
        frame = Frame(margin_x, start_y - 12 * cm, page_width - 2 * margin_x, 6 * cm, showBoundary=False)
        story = [Paragraph(text, text_style)]
        frame.addFromList(story, c)

        # Informações adicionais
        y_position = start_y - 14 * cm
        c.setFont("Helvetica", 10)
        c.drawString(margin_x, y_position, "Tamanho     [     ] P               [     ] M               [     ] G")
        c.drawString(margin_x, y_position - 1 * cm, "Número de Série ____________________")
        c.drawString(margin_x, y_position - 2 * cm, "Data: ________/________/________")

        # Campos de assinatura com verificação do limite inferior
        signature_y_position = y_position - 5 * cm
        for line, label in enumerate([
            "Assinatura Policial Penal: ______________________________________",
            "CPF: ______________________________________",
            "Assinatura Chefe Imediato: ______________________________________",
            "Matrícula: ______________________________________",
            "Assinatura Membro da Comissão: ______________________________________",
            "Matrícula: ______________________________________"
        ]):
            # Limite de quebra na borda inferior
            if signature_y_position - (line * cm) > 2 * cm:
                c.drawString(margin_x, signature_y_position - (line * cm), label)

        # Finalização do PDF
        c.showPage()
        c.save()
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()

    def generate_zip(self, servidores):
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for servidor in servidores:
                pdf_content = self.generate_pdf(servidor)
                if pdf_content:
                    # Adiciona o PDF gerado ao arquivo ZIP
                    zip_file.writestr(f"termo_acautelamento_{servidor.nome}.pdf", pdf_content)

        zip_buffer.seek(0)
        return zip_buffer

    def get(self, request, *args, **kwargs):
        if 'download_zip' in request.GET:
            servidores = self.get_queryset()

            # Verifica se a quantidade de servidores excede o limite
            if servidores.count() > self.max_pdfs:
                messages.error(request, f"O limite máximo de geração é de {self.max_pdfs} PDFs por vez.")
                return redirect('gesipe:gesipe_armaria')

            if not servidores.exists():
                messages.error(request, "Nenhum registro encontrado para os filtros selecionados.")
                return redirect('gesipe:gesipe_armaria')

            zip_buffer = self.generate_zip(servidores)
            response = HttpResponse(zip_buffer, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="termos_acautelamento.zip"'
            return response
        else:
            return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        # Lógica para limitar a exibição das páginas
        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)
        context['generos'] = Servidor.objects.values_list('genero', flat=True).distinct()
        context['cargos'] = Servidor.objects.values_list('cargo', flat=True).distinct()
        context['cargos_comissionado'] = [cargo for cargo in
                                          Servidor.objects.values_list('cargo_comissionado', flat=True).distinct()]
        return context




class Gesipe(LoginRequiredMixin, ListView):
    template_name = "gesipe.html"
    model = Gesipe_adm
    paginate_by = 20
    ordering = ['-data']

    def get_queryset(self):
        query = self.request.GET.get('query', '')
        queryset = super().get_queryset()

        if query:
            # Tentar interpretar 'query' como data nos formatos DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD
            data_pesquisa = self.parse_data(query)
            if data_pesquisa:
                queryset = queryset.filter(data=data_pesquisa)
            else:
                # Pesquisa normal por outros campos se não for data
                queryset = queryset.filter(
                    Q(usuario__username__icontains=query) |
                    Q(total__icontains=query)
                )

        return queryset

    def post(self, request, *args, **kwargs):
        data_pesquisa = request.POST.get('data_pesquisa')

        if data_pesquisa:
            # Tentar interpretar a data nos formatos permitidos
            data_pesquisa = self.parse_data(data_pesquisa)
            if not data_pesquisa:
                return JsonResponse({'error': 'Data inválida'}, status=400)

            if Gesipe_adm.objects.filter(data=data_pesquisa).exists():
                existing_record = Gesipe_adm.objects.get(data=data_pesquisa)
                edit_url = reverse('gesipe:gesipe_adm_edit', kwargs={'pk': existing_record.pk})
                return JsonResponse({'exists': True, 'url': edit_url})
            else:
                create_url = reverse('gesipe:gesipe_adm')
                return JsonResponse({'exists': False, 'url': create_url})

        return JsonResponse({'error': 'Data não fornecida'}, status=400)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('query', '')

        # total de movimentacoes gesipe admin
        context['total_mov_adm'] = Gesipe_adm.objects.count()

        # Obtendo dados gerais para o gráfico de linha
        daily_totals = Gesipe_adm.objects.values('data') \
            .annotate(total=Sum('total')) \
            .order_by('data')

        # Separar labels e valores
        line_labels = [entry['data'].strftime("%d/%m/%Y") for entry in daily_totals]
        line_values = [entry['total'] for entry in daily_totals]

        # Passar os dados para o contexto
        context['line_labels'] = line_labels
        context['line_values'] = line_values

        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        # Lógica para limitar a exibição das páginas
        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)

        # Obtendo os dados para o gráfico de barra administrativo
        current_year = timezone.now().year
        monthly_totals_adm_barra = []

        for month in range(1, 13):
            monthly_total = \
                Gesipe_adm.objects.filter(data__year=current_year, data__month=month).aggregate(total=Sum('total'))[
                    'total'] or 0
            monthly_totals_adm_barra.append(monthly_total)

        # Labels dos meses
        # Dicionário com os nomes dos meses em português
        meses_em_portugues = {
            1: "Janeiro",
            2: "Fevereiro",
            3: "Março",
            4: "Abril",
            5: "Maio",
            6: "Junho",
            7: "Julho",
            8: "Agosto",
            9: "Setembro",
            10: "Outubro",
            11: "Novembro",
            12: "Dezembro"
        }
        labels_meses = [meses_em_portugues[month] for month in range(1, 13)]

        # Obtendo os dados para o gráfico de pizza administrativo
        total_values = Gesipe_adm.objects.aggregate(
            diarias=Sum('diarias'),
            documentos_capturados=Sum('documentos_capturados'),
            total_despacho=Sum('total_despacho'),
            total_oficio=Sum('total_oficio'),
            total_os=Sum('total_os'),
            processos=Sum('processos'),
            portarias=Sum('portarias')
        )

        pie_labels_adm = [
            'Diárias',
            'Documentos Capturados',
            'Despachos',
            'Ofícios',
            'OS',
            'Processos',
            'Portarias'
        ]
        pie_values_adm = [
            total_values['diarias'] or 0,
            total_values['documentos_capturados'] or 0,
            total_values['total_despacho'] or 0,
            total_values['total_oficio'] or 0,
            total_values['total_os'] or 0,
            total_values['processos'] or 0,
            total_values['portarias'] or 0,
        ]

        # Passando os dados para o template
        context['labels_mensais'] = labels_meses
        context['values_mensais'] = monthly_totals_adm_barra
        context['pie_labels_adm'] = pie_labels_adm
        context['pie_values_adm'] = pie_values_adm

        return context

    def parse_data(self, date_str):
        """
        Tenta fazer o parsing de uma string de data nos formatos DD/MM/YYYY, DD-MM-YYYY e YYYY-MM-DD.
        Retorna a data no formato 'datetime.date' se válido, senão retorna None.
        """
        date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None


class GesipeAdm(LoginRequiredMixin, CreateView):
    model = Gesipe_adm
    form_class = GesipeAdmForm
    template_name = 'gesipe_adm.html'
    success_url = reverse_lazy('gesipe:gesipe_adm')

    def form_valid(self, form):
        # Se o formulário passar pela validação, salva o novo registro
        print("Formulário válido, salvando dados...")
        dados_adm = form.save(commit=False)
        dados_adm.usuario = self.request.user  # Atribui o usuário atual
        dados_adm.data_edicao = timezone.now()  # Define a data de edição
        dados_adm.save()
        messages.success(self.request, 'Dados adicionados com sucesso!')
        return redirect(self.success_url)

    def form_invalid(self, form):

        if 'data' in form.errors:
            # Exibe uma mensagem de erro personalizada

            messages.error(self.request, 'Registro para a data já existe. Utilize a opção de edição para alterar os dados.')
        else:
            pass

        # Retorna o formulário inválido para o template com a mensagem de erro
        return self.render_to_response(self.get_context_data(form=form))


class GesipeAdmEdit(LoginRequiredMixin, UpdateView):
    model = Gesipe_adm
    form_class = GesipeAdmForm
    template_name = 'gesipe_adm_edit.html'
    success_url = reverse_lazy('gesipe:gesipe_adm')  # Redirecionar para a página de sucesso

    def get_form_kwargs(self):
        """Passa o formato correto da data para o formulário."""
        kwargs = super().get_form_kwargs()
        if self.object:
            # Ajusta o valor da data para o formato 'yyyy-mm-dd'
            kwargs['initial'] = {
                'data': self.object.data.strftime('%Y-%m-%d'),
            }
        return kwargs

    def form_valid(self, form):
        if self.request.POST.get('action') == 'delete':
            # Lida com a exclusão do registro
            self.object.delete()
            messages.error(self.request, 'Registro excluído com sucesso!')
            return redirect(self.success_url)

        # Atualiza os dados do registro
        dados_adm = form.save(commit=False)
        dados_adm.data_edicao = timezone.now()  # Atualiza a data de edição
        dados_adm.save()

        messages.success(self.request, 'Dados atualizados com sucesso!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formatted_date'] = self.object.data.strftime('%d/%m/%Y')
        return context



class GesipeAdmLote(LoginRequiredMixin, View):
    form_class = UploadFileForm
    template_name = 'gesipe_adm_lote.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            arquivo_excel = request.FILES['arquivo_excel']

            # Carregar a planilha com valores calculados em vez de fórmulas
            wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
            sheet = wb['Planilha']  # Nome correto da aba

            try:
                for row in sheet.iter_rows(min_row=4, values_only=True):
                    if not any(row):  # Ignorar linhas em branco
                        continue

                    # Extrair o valor da data
                    data = row[0]  # Supondo que a data está na primeira coluna


                    # Converter a data para o formato 'YYYY-MM-DD' se necessário
                    if isinstance(data, str):  # Se a data estiver como string
                        try:
                            data = datetime.strptime(data, '%d/%m/%Y').date()  # Converter de DD/MM/YYYY para objeto date
                        except ValueError:
                            messages.error(request, f'O formato de data "{data}" é inválido. Use o formato DD/MM/YYYY.')
                            return render(request, self.template_name, {'form': form})


                    # Se 'created' for True, significa que o registro foi criado; se False, ele foi atualizado.
                    obj, created = Gesipe_adm.objects.update_or_create(
                        data=data,  # Critério de identificação pelo campo `data`
                        defaults={
                            'processos': 0 if row[1] is None else row[1],
                            'memorandos_diarias': 0 if row[2] is None else row[2],
                            'memorandos_documentos_capturados': 0 if row[3] is None else row[3],
                            'despachos_gerencias': 0 if row[4] is None else row[4],
                            'despachos_unidades': 0 if row[5] is None else row[5],
                            'despachos_grupos': 0 if row[6] is None else row[6],
                            'oficios_internos_unidades_prisionais': 0 if row[7] is None else row[7],
                            'oficios_internos_setores_seap_pb': 0 if row[8] is None else row[8],
                            'oficios_internos_circular': 0 if row[9] is None else row[9],
                            'oficios_externos_seap_pb': 0 if row[10] is None else row[10],
                            'oficios_externos_diversos': 0 if row[11] is None else row[11],
                            'oficios_externos_judiciario': 0 if row[12] is None else row[12],
                            'os_grupos': 0 if row[13] is None else row[13],
                            'os_diversos': 0 if row[14] is None else row[14],
                            'portarias': 0 if row[15] is None else row[15],
                            'usuario': request.user  # Assumindo que você quer armazenar o usuário que fez a alteração
                        }
                    )


                messages.success(request, 'Dados importados com sucesso!')
            except Exception as e:
                messages.error(request, f'Ocorreu um erro ao processar o arquivo: {e}')

            return redirect('gesipe:gesipe_adm_lote')

        return render(request, self.template_name, {'form': form})


class RelatorioGesipeAdm(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Gesipe_adm
    template_name = "relatorio_gesipe_adm.html"
    paginate_by = 20
    ordering = ['-data']

    def test_func(self):
        user = self.request.user
        grupos_permitidos = ['Administrador', 'GerGesipe', 'ServGesipe']
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)

    def get_queryset(self, request=None):
        queryset = super().get_queryset()

        # Se o request for passado, filtra o queryset com base nas datas
        if request:
            data_inicial = request.GET.get('dataInicial')
            data_final = request.GET.get('dataFinal')
            if data_inicial and data_final:
                try:
                    data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
                    data_final = datetime.strptime(data_final, '%Y-%m-%d').date()
                    queryset = queryset.filter(data__range=(data_inicial, data_final))
                except ValueError:
                    pass  # Ignora erros de formato de data

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['dataInicial'] = self.request.GET.get('dataInicial', '')
        context['dataFinal'] = self.request.GET.get('dataFinal', '')
        context['setor'] = 'Administrativo'

        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        # Limite de páginas
        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)

        return context

    def get(self, request, *args, **kwargs):
        action = request.GET.get('action')
        if action == 'export_pdf':
            return export_pdf(request)

        return super().get(request, *args, **kwargs)


#GESIPE SGA

class GesipeSga(LoginRequiredMixin, ListView):
    template_name = "gesipe_sga.html"
    model = Gesipe_adm
    paginate_by = 10
    ordering = ['-data']

    def get_queryset(self):
        query = self.request.GET.get('query', '')
        queryset = super().get_queryset()

        if query:
            # Tentar interpretar 'query' como data nos formatos DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD
            data_pesquisa = self.parse_data(query)
            if data_pesquisa:
                queryset = queryset.filter(data=data_pesquisa)
            else:
                # Pesquisa normal por outros campos se não for data
                queryset = queryset.filter(
                    Q(usuario__username__icontains=query) |
                    Q(total__icontains=query)
                )

        return queryset

    def post(self, request, *args, **kwargs):
        data_pesquisa = request.POST.get('data_pesquisa')

        if data_pesquisa:
            # Tentar interpretar a data nos formatos permitidos
            data_pesquisa = self.parse_data(data_pesquisa)
            if not data_pesquisa:
                return JsonResponse({'error': 'Data inválida'}, status=400)

            if Gesipe_adm.objects.filter(data=data_pesquisa).exists():
                existing_record = Gesipe_adm.objects.get(data=data_pesquisa)
                edit_url = reverse('gesipe:gesipe_adm_edit', kwargs={'pk': existing_record.pk})
                return JsonResponse({'exists': True, 'url': edit_url})
            else:
                create_url = reverse('gesipe:gesipe_adm')
                return JsonResponse({'exists': False, 'url': create_url})

        return JsonResponse({'error': 'Data não fornecida'}, status=400)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('query', '')

        # total de movimentacoes gesipe admin
        context['total_mov_adm'] = Gesipe_adm.objects.count()

        # Obtendo dados gerais para o gráfico de linha
        daily_totals = Gesipe_adm.objects.values('data') \
            .annotate(total=Sum('total')) \
            .order_by('data')

        # Separar labels e valores
        line_labels = [entry['data'].strftime("%d/%m/%Y") for entry in daily_totals]
        line_values = [entry['total'] for entry in daily_totals]

        # Passar os dados para o contexto
        context['line_labels'] = line_labels
        context['line_values'] = line_values

        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        # Lógica para limitar a exibição das páginas
        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)

        # Obtendo os dados para o gráfico de barra administrativo
        current_year = timezone.now().year
        monthly_totals_adm_barra = []

        for month in range(1, 13):
            monthly_total = \
            Gesipe_adm.objects.filter(data__year=current_year, data__month=month).aggregate(total=Sum('total'))[
                'total'] or 0
            monthly_totals_adm_barra.append(monthly_total)

        # Labels dos meses
        # Dicionário com os nomes dos meses em português
        meses_em_portugues = {
            1: "Janeiro",
            2: "Fevereiro",
            3: "Março",
            4: "Abril",
            5: "Maio",
            6: "Junho",
            7: "Julho",
            8: "Agosto",
            9: "Setembro",
            10: "Outubro",
            11: "Novembro",
            12: "Dezembro"
        }
        labels_meses = [meses_em_portugues[month] for month in range(1, 13)]

        # Obtendo os dados para o gráfico de pizza administrativo
        total_values = Gesipe_adm.objects.aggregate(
            diarias=Sum('diarias'),
            documentos_capturados=Sum('documentos_capturados'),
            total_despacho=Sum('total_despacho'),
            total_oficio=Sum('total_oficio'),
            total_os=Sum('total_os'),
            processos=Sum('processos'),
            portarias=Sum('portarias')
        )

        pie_labels_adm = [
            'Diárias',
            'Documentos Capturados',
            'Despachos',
            'Ofícios',
            'OS',
            'Processos',
            'Portarias'
        ]
        pie_values_adm = [
            total_values['diarias'] or 0,
            total_values['documentos_capturados'] or 0,
            total_values['total_despacho'] or 0,
            total_values['total_oficio'] or 0,
            total_values['total_os'] or 0,
            total_values['processos'] or 0,
            total_values['portarias'] or 0,
        ]

        # Passando os dados para o template
        context['labels_mensais'] = labels_meses
        context['values_mensais'] = monthly_totals_adm_barra
        context['pie_labels_adm'] = pie_labels_adm
        context['pie_values_adm'] = pie_values_adm

        return context

    def parse_data(self, date_str):
        """
        Tenta fazer o parsing de uma string de data nos formatos DD/MM/YYYY, DD-MM-YYYY e YYYY-MM-DD.
        Retorna a data no formato 'datetime.date' se válido, senão retorna None.
        """
        date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None


# Gesipe SGA




class GesipeSga(LoginRequiredMixin, ListView):
    template_name = "sga/gesipe_sga.html"
    model = Gesipe_Sga
    paginate_by = 20
    ordering = ['-data']

    def get_queryset(self):
        query = self.request.GET.get('query', '')
        queryset = super().get_queryset()

        if query:
            # Tentar interpretar 'query' como data nos formatos DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD
            data_pesquisa = self.parse_data(query)
            if data_pesquisa:
                queryset = queryset.filter(data=data_pesquisa)
            else:
                # Pesquisa normal por outros campos se não for data
                queryset = queryset.filter(
                    Q(usuario__username__icontains=query) |
                    Q(total__icontains=query)
                )

        return queryset

    def post(self, request, *args, **kwargs):
        data_pesquisa = request.POST.get('data_pesquisa')

        if data_pesquisa:
            # Tentar interpretar a data nos formatos permitidos
            data_pesquisa = self.parse_data(data_pesquisa)
            if not data_pesquisa:
                return JsonResponse({'error': 'Data inválida'}, status=400)

            if Gesipe_Sga.objects.filter(data=data_pesquisa).exists():
                existing_record = Gesipe_Sga.objects.get(data=data_pesquisa)
                edit_url = reverse('gesipe:gesipe_sga_edit', kwargs={'pk': existing_record.pk})
                return JsonResponse({'exists': True, 'url': edit_url})
            else:
                create_url = reverse('gesipe:gesipe_sga_add')
                return JsonResponse({'exists': False, 'url': create_url})

        return JsonResponse({'error': 'Data não fornecida'}, status=400)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('query', '')

        # Total de movimentações Gesipe SGA
        context['total_mov_sga'] = Gesipe_Sga.objects.count()

        # Obtendo dados gerais para o gráfico de linha
        daily_totals = (
            Gesipe_Sga.objects.values('data')
            .annotate(
                total=(
                    Sum(
                        F('agendamentos_entradas') +
                        F('comunicacoes_presos') +
                        F('comunicacoes_servidores') +
                        F('comunicacoes_setores') +
                        F('comunicacoes_judiciais_externas') +
                        F('om_grupos') +
                        F('om_unidades')
                    )
                )
            )
            .order_by('data')
        )

        # Separar labels e valores para o gráfico de linha
        line_labels = [entry['data'].strftime("%d/%m/%Y") for entry in daily_totals]
        line_values = [entry['total'] for entry in daily_totals]

        context['line_labels'] = line_labels
        context['line_values'] = line_values

        # Lógica de paginação
        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)

        # Obtendo os dados para o gráfico de barra SGA
        current_year = timezone.now().year
        monthly_totals_sga_barra = []

        for month in range(1, 13):
            monthly_total = (
                    Gesipe_Sga.objects.filter(data__year=current_year, data__month=month)
                    .aggregate(
                        total=Sum(
                            F('agendamentos_entradas') +
                            F('comunicacoes_presos') +
                            F('comunicacoes_servidores') +
                            F('comunicacoes_setores') +
                            F('comunicacoes_judiciais_externas') +
                            F('om_grupos') +
                            F('om_unidades')
                        )
                    )['total']
                    or 0
            )
            monthly_totals_sga_barra.append(monthly_total)


        # Labels dos meses
        meses_em_portugues = {
            1: "Janeiro",
            2: "Fevereiro",
            3: "Março",
            4: "Abril",
            5: "Maio",
            6: "Junho",
            7: "Julho",
            8: "Agosto",
            9: "Setembro",
            10: "Outubro",
            11: "Novembro",
            12: "Dezembro",
        }
        labels_meses = [meses_em_portugues[month] for month in range(1, 13)]

        # Passando os dados para o template
        context['labels_mensais'] = labels_meses
        context['values_mensais'] = monthly_totals_sga_barra

        # Obtendo os dados para o gráfico de pizza administrativo
        total_values = Gesipe_Sga.objects.aggregate(
            total_agendamentos_entradas=Sum('agendamentos_entradas'),
            total_comunicacoes_presos=Sum('comunicacoes_presos'),
            total_comunicacoes_servidores=Sum('comunicacoes_servidores'),
            total_comunicacoes_setores=Sum('comunicacoes_setores'),
            total_comunicacoes_judiciais_externas=Sum('comunicacoes_judiciais_externas'),
            total_om_grupos=Sum('om_grupos'),
            total_om_unidades=Sum('om_unidades'),
        )

        pie_labels_sga = [
            'Agendamentos Entradas',
            'Comunicações Presos',
            'Comunicações Servidores',
            'Comunicações Setores',
            'Comunicações Judiciais e Externas',
            'Ordens de Missão Grupos',
            'Ordens de Missão unidades',
        ]
        pie_values_sga = [
            total_values['total_agendamentos_entradas'] or 0,
            total_values['total_comunicacoes_presos'] or 0,
            total_values['total_comunicacoes_servidores'] or 0,
            total_values['total_comunicacoes_setores'] or 0,
            total_values['total_comunicacoes_judiciais_externas'] or 0,
            total_values['total_om_grupos'] or 0,
            total_values['total_om_unidades'] or 0,
        ]

        context['pie_labels_sga'] = pie_labels_sga
        context['pie_values_sga'] = pie_values_sga

        return context

    def parse_data(self, date_str):
        """
        Tenta fazer o parsing de uma string de data nos formatos DD/MM/YYYY, DD-MM-YYYY e YYYY-MM-DD.
        Retorna a data no formato 'datetime.date' se válido, senão retorna None.
        """
        date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None


class GesipeSgaAdd(LoginRequiredMixin, CreateView):
    model = Gesipe_Sga
    form_class = GesipeSgaForm
    template_name = 'sga/gesipe_sga_add.html'
    success_url = reverse_lazy('gesipe:gesipe_sga_add')

    def form_valid(self, form):
        # Se o formulário passar pela validação, salva o novo registro

        dados_sga = form.save(commit=False)
        dados_sga.usuario = self.request.user  # Atribui o usuário atual
        dados_sga.data_edicao = timezone.now()  # Define a data de edição
        dados_sga.save()
        messages.success(self.request, 'Dados adicionados com sucesso!')
        return redirect(self.success_url)


class GesipeSgaEdit(LoginRequiredMixin, UpdateView):
    model = Gesipe_Sga
    form_class = GesipeSgaForm
    template_name = 'sga/gesipe_sga_edit.html'
    success_url = reverse_lazy('gesipe:gesipe_sga_add')  # Redirecionar para a página de sucesso

    def get_form_kwargs(self):
        """Passa o formato correto da data para o formulário."""
        kwargs = super().get_form_kwargs()
        if self.object:
            # Ajusta o valor da data para o formato 'yyyy-mm-dd'
            kwargs['initial'] = {
                'data': self.object.data.strftime('%Y-%m-%d'),
            }
        return kwargs

    def form_valid(self, form):
        if self.request.POST.get('action') == 'delete':
            # Lida com a exclusão do registro
            self.object.delete()
            messages.error(self.request, 'Registro excluído com sucesso!')
            return redirect(self.success_url)

        # Atualiza os dados do registro
        dados_sga = form.save(commit=False)
        dados_sga.data_edicao = timezone.now()  # Atualiza a data de edição
        dados_sga.save()

        messages.success(self.request, 'Dados atualizados com sucesso!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formatted_date'] = self.object.data.strftime('%d/%m/%Y')
        return context


class RelatorioGesipeSga(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Gesipe_Sga
    template_name = "sga/relatorio_gesipe_sga.html"
    paginate_by = 20
    ordering = ['-data']

    def test_func(self):
        user = self.request.user
        grupos_permitidos = ['Administrador', 'GerGesipe', 'ServGesipeSga']
        return user.groups.filter(name__in=grupos_permitidos).exists()

    def handle_no_permission(self):
        messages.error(self.request, "Você não tem permissão para acessar esta página.")
        return render(self.request, '403.html', status=403)

    def get_queryset(self, request=None):
        queryset = super().get_queryset()

        # Se o request for passado, filtra o queryset com base nas datas
        if request:
            data_inicial = request.GET.get('dataInicial')
            data_final = request.GET.get('dataFinal')
            if data_inicial and data_final:
                try:
                    data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
                    data_final = datetime.strptime(data_final, '%Y-%m-%d').date()
                    queryset = queryset.filter(data__range=(data_inicial, data_final))
                except ValueError:
                    pass  # Ignora erros de formato de data

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['dataInicial'] = self.request.GET.get('dataInicial', '')
        context['dataFinal'] = self.request.GET.get('dataFinal', '')

        page_obj = context['page_obj']
        paginator = page_obj.paginator
        page_range = paginator.page_range

        # Limite de páginas
        if page_obj.number > 3:
            start = page_obj.number - 2
        else:
            start = 1

        if page_obj.number < paginator.num_pages - 2:
            end = page_obj.number + 2
        else:
            end = paginator.num_pages

        context['page_range'] = range(start, end + 1)

        return context

    def get(self, request, *args, **kwargs):
        action = request.GET.get('action')
        setor = "Sga"
        if action == 'export_pdf_sga_detalhado':
            return export_pdf_sga_detalhado(request, setor)

        return super().get(request, *args, **kwargs)



class GesipeSgaLote(LoginRequiredMixin, View):
    form_class = UploadFileForm
    template_name = 'sga/gesipe_sga_lote.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            arquivo_excel = request.FILES['arquivo_excel']

            # Carregar a planilha com valores calculados em vez de fórmulas
            wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
            sheet = wb['Planilha']  # Nome correto da aba

            try:
                for row in sheet.iter_rows(min_row=4, values_only=True):
                    if not any(row):  # Ignorar linhas em branco
                        continue

                    # Extrair o valor da data
                    data = row[0]  # Supondo que a data está na primeira coluna

                    # Converter a data para o formato 'YYYY-MM-DD' se necessário
                    if isinstance(data, str):  # Se a data estiver como string
                        try:
                            data = datetime.strptime(data,
                                                     '%d/%m/%Y').date()  # Converter de DD/MM/YYYY para objeto date
                        except ValueError:
                            messages.error(request, f'O formato de data "{data}" é inválido. Use o formato DD/MM/YYYY.')
                            return render(request, self.template_name, {'form': form})

                    # Excluir todos os registros existentes com a mesma data
                    Gesipe_Sga.objects.filter(data=data).delete()

                    # Criar um novo registro
                    Gesipe_Sga.objects.create(
                        data=data,
                        agendamentos_entradas=0 if row[1] is None else row[1],
                        comunicacoes_presos=0 if row[2] is None else row[2],
                        comunicacoes_servidores=0 if row[3] is None else row[3],
                        comunicacoes_setores=0 if row[4] is None else row[4],
                        comunicacoes_judiciais_externas=0 if row[5] is None else row[5],
                        om_grupos=0 if row[6] is None else row[6],
                        om_unidades=0 if row[7] is None else row[7],
                        data_edicao=timezone.now(),  # Define a data de edição
                        usuario=request.user  # Assumindo que você quer armazenar o usuário que fez a alteração
                    )

                messages.success(request, 'Dados importados com sucesso!')
            except Exception as e:
                messages.error(request, f'Ocorreu um erro ao processar o arquivo: {e}')
                return render(request, self.template_name, {'form': form})

            return redirect('gesipe:gesipe_sga_lote')

        return render(request, self.template_name, {'form': form})
